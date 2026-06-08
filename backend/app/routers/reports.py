"""Raporlar: Satis ziyaret kayitlarini listeleme, filtreleme ve disa aktarma."""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from datetime import date, datetime, timedelta
from io import BytesIO

from ..database import get_db
from ..models import SalesVisit, Customer, User
from ..auth import get_current_user, require_admin, SECRET_KEY, ALGORITHM
from jose import jwt

router = APIRouter(prefix="/api/reports", tags=["reports"])


def _get_user_from_token(token: str, db: Session) -> User:
    """Query parameter token'dan kullanici al (download linkleri icin)."""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = int(payload.get("sub"))
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(401, "Gecersiz token")
        return user
    except Exception:
        raise HTTPException(401, "Gecersiz token")


def _parse_customer_ids(customer_id, customer_ids):
    """Tekli customer_id veya virgulle ayrilmis customer_ids'i ID listesine cevirir."""
    ids = []
    if customer_ids:
        ids = [int(x) for x in str(customer_ids).split(",") if x.strip().isdigit()]
    elif customer_id:
        ids = [customer_id]
    return ids


def _apply_date_filter(q, start_date, end_date, period):
    """Tarih filtresini uygula."""
    if period == "today":
        start_date = date.today()
        end_date = date.today()
    elif period == "week":
        start_date = date.today() - timedelta(days=date.today().weekday())
        end_date = date.today()
    elif period == "month":
        start_date = date.today().replace(day=1)
        end_date = date.today()

    if start_date:
        q = q.filter(SalesVisit.visit_date >= start_date)
    if end_date:
        q = q.filter(SalesVisit.visit_date <= end_date)
    return q, start_date, end_date


# ── Satis Kayitlarini Listele ──
@router.get("/sales")
def list_sales(
    start_date: date | None = None,
    end_date: date | None = None,
    period: str | None = None,
    user_id: int | None = None,
    customer_id: int | None = None,
    customer_ids: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(SalesVisit).options(
        joinedload(SalesVisit.user),
        joinedload(SalesVisit.customer),
    )

    q, start_date, end_date = _apply_date_filter(q, start_date, end_date, period)

    if user_id:
        q = q.filter(SalesVisit.user_id == user_id)
    cust_ids = _parse_customer_ids(customer_id, customer_ids)
    if cust_ids:
        q = q.filter(SalesVisit.customer_id.in_(cust_ids))

    # Admin degil ise sadece kendi kayitlari
    if current_user.role != "admin":
        q = q.filter(SalesVisit.user_id == current_user.id)

    visits = q.order_by(SalesVisit.visit_date.desc(), SalesVisit.id.desc()).all()

    result = []
    for v in visits:
        result.append({
            "id": v.id,
            "user_id": v.user_id,
            "user_name": v.user.full_name if v.user else "-",
            "customer_id": v.customer_id,
            "customer_name": v.customer.name if v.customer else "-",
            "visit_date": str(v.visit_date) if v.visit_date else None,
            "sale_amount": v.sale_amount or 0,
            "visited": v.visited,
            "notes": v.notes,
            "check_in_at": str(v.check_in_at) if v.check_in_at else None,
            "check_out_at": str(v.check_out_at) if v.check_out_at else None,
        })
    return result


# ── Ozet Istatistikler ──
@router.get("/summary")
def report_summary(
    start_date: date | None = None,
    end_date: date | None = None,
    period: str | None = None,
    user_id: int | None = None,
    customer_id: int | None = None,
    customer_ids: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(SalesVisit)
    q, start_date, end_date = _apply_date_filter(q, start_date, end_date, period)

    if user_id:
        q = q.filter(SalesVisit.user_id == user_id)
    cust_ids = _parse_customer_ids(customer_id, customer_ids)
    if cust_ids:
        q = q.filter(SalesVisit.customer_id.in_(cust_ids))
    if current_user.role != "admin":
        q = q.filter(SalesVisit.user_id == current_user.id)

    visits = q.all()

    total_records = len(visits)
    total_revenue = sum(v.sale_amount or 0 for v in visits)
    visited_count = sum(1 for v in visits if v.visited == 1)
    not_visited_count = total_records - visited_count
    avg_sale = total_revenue / visited_count if visited_count > 0 else 0

    # ST bazli ozet
    rep_stats = {}
    for v in visits:
        uid = v.user_id
        if uid not in rep_stats:
            rep_stats[uid] = {"user_name": "", "total_sales": 0, "visit_count": 0}
        if v.user:
            rep_stats[uid]["user_name"] = v.user.full_name
        rep_stats[uid]["total_sales"] += v.sale_amount or 0
        if v.visited == 1:
            rep_stats[uid]["visit_count"] += 1

    return {
        "total_records": total_records,
        "total_revenue": round(total_revenue, 2),
        "visited_count": visited_count,
        "not_visited_count": not_visited_count,
        "avg_sale": round(avg_sale, 2),
        "rep_stats": list(rep_stats.values()),
    }


# ── Excel Disa Aktarma ──
@router.get("/export/excel")
def export_excel(
    token: str = Query(...),
    start_date: date | None = None,
    end_date: date | None = None,
    period: str | None = None,
    user_id: int | None = None,
    customer_id: int | None = None,
    customer_ids: str | None = None,
    db: Session = Depends(get_db),
):
    current_user = _get_user_from_token(token, db)

    q = db.query(SalesVisit).options(
        joinedload(SalesVisit.user),
        joinedload(SalesVisit.customer),
    )
    q, start_date, end_date = _apply_date_filter(q, start_date, end_date, period)

    if user_id:
        q = q.filter(SalesVisit.user_id == user_id)
    cust_ids = _parse_customer_ids(customer_id, customer_ids)
    if cust_ids:
        q = q.filter(SalesVisit.customer_id.in_(cust_ids))
    if current_user.role != "admin":
        q = q.filter(SalesVisit.user_id == current_user.id)

    visits = q.order_by(SalesVisit.visit_date.desc(), SalesVisit.id.desc()).all()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Satis Raporu"

    # Baslik
    ws.merge_cells("A1:G1")
    ws["A1"] = "Satis Raporu"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Tarih bilgisi
    ws.merge_cells("A2:G2")
    tarih_text = "Tum Tarihler"
    if start_date and end_date:
        tarih_text = f"{start_date} - {end_date}"
    elif start_date:
        tarih_text = f"{start_date} itibariyle"
    ws["A2"] = f"Donem: {tarih_text} | Toplam: {len(visits)} kayit"
    ws["A2"].font = Font(size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    # Basliklar
    headers = ["#", "Tarih", "Satis Temsilcisi", "Musteri", "Satis Tutari (TL)", "Ziyaret", "Notlar"]
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    header_font = Font(bold=True, size=10, color="334155")
    thin_border = Border(
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[3].height = 28

    # Veri
    total_sales = 0
    for i, v in enumerate(visits, 1):
        row = i + 3
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=str(v.visit_date) if v.visit_date else "-")
        ws.cell(row=row, column=3, value=v.user.full_name if v.user else "-")
        ws.cell(row=row, column=4, value=v.customer.name if v.customer else "-")
        amount = v.sale_amount or 0
        total_sales += amount
        amt_cell = ws.cell(row=row, column=5, value=amount)
        amt_cell.number_format = '#,##0.00'
        ws.cell(row=row, column=6, value="Evet" if v.visited == 1 else "Hayir")
        ws.cell(row=row, column=7, value=v.notes or "-")

        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = Alignment(vertical="center")

    # Toplam satiri
    total_row = len(visits) + 4
    ws.merge_cells(f"A{total_row}:D{total_row}")
    ws.cell(row=total_row, column=1, value="TOPLAM").font = Font(bold=True, size=11)
    total_cell = ws.cell(row=total_row, column=5, value=total_sales)
    total_cell.font = Font(bold=True, size=11)
    total_cell.number_format = '#,##0.00'
    ws.cell(row=total_row, column=6, value=f"{sum(1 for v in visits if v.visited==1)} ziyaret")

    # Sutun genislikleri
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 40

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=satis_raporu.xlsx"},
    )


# ── PDF Disa Aktarma ──
@router.get("/export/pdf")
def export_pdf(
    token: str = Query(...),
    start_date: date | None = None,
    end_date: date | None = None,
    period: str | None = None,
    user_id: int | None = None,
    customer_id: int | None = None,
    customer_ids: str | None = None,
    db: Session = Depends(get_db),
):
    current_user = _get_user_from_token(token, db)

    q = db.query(SalesVisit).options(
        joinedload(SalesVisit.user),
        joinedload(SalesVisit.customer),
    )
    q, start_date, end_date = _apply_date_filter(q, start_date, end_date, period)

    if user_id:
        q = q.filter(SalesVisit.user_id == user_id)
    cust_ids = _parse_customer_ids(customer_id, customer_ids)
    if cust_ids:
        q = q.filter(SalesVisit.customer_id.in_(cust_ids))
    if current_user.role != "admin":
        q = q.filter(SalesVisit.user_id == current_user.id)

    visits = q.order_by(SalesVisit.visit_date.desc(), SalesVisit.id.desc()).all()

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = []

    # Baslik
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#1e3a5f"))
    story.append(Paragraph("Satis Raporu", title_style))
    story.append(Spacer(1, 8))

    tarih_text = "Tum Tarihler"
    if start_date and end_date:
        tarih_text = f"{start_date} - {end_date}"
    info_style = ParagraphStyle("info", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
    story.append(Paragraph(f"Donem: {tarih_text}  |  Toplam: {len(visits)} kayit  |  Olusturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}", info_style))
    story.append(Spacer(1, 14))

    # Ozet
    total_sales = sum(v.sale_amount or 0 for v in visits)
    visited = sum(1 for v in visits if v.visited == 1)
    avg = total_sales / visited if visited > 0 else 0
    summary_data = [
        ["Toplam Kayit", "Ziyaret Edilen", "Toplam Satis", "Ortalama Satis"],
        [str(len(visits)), str(visited), f"{total_sales:,.0f} TL", f"{avg:,.0f} TL"],
    ]
    summary_table = Table(summary_data, colWidths=[6*cm, 5*cm, 6*cm, 6*cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, 1), 12),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 16))

    # Detay tablosu
    note_style = ParagraphStyle("note", parent=styles["Normal"], fontSize=7, leading=9)
    header = ["#", "Tarih", "Satis Temsilcisi", "Musteri", "Tutar (TL)", "Ziyaret", "Notlar"]
    data = [header]
    for i, v in enumerate(visits, 1):
        notes_text = (v.notes or "-")[:80]
        data.append([
            str(i),
            str(v.visit_date) if v.visit_date else "-",
            v.user.full_name if v.user else "-",
            v.customer.name if v.customer else "-",
            f"{v.sale_amount or 0:,.0f}",
            "Evet" if v.visited == 1 else "Hayir",
            Paragraph(notes_text, note_style),
        ])

    col_widths = [1.2*cm, 2.8*cm, 4.5*cm, 5.5*cm, 3*cm, 2*cm, 7*cm]
    detail_table = Table(data, colWidths=col_widths, repeatRows=1)
    detail_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (6, 1), (6, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(detail_table)

    doc.build(story)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=satis_raporu.pdf"},
    )
